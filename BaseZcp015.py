import pandas as pd
from time import sleep
import os
import openpyxl as xl
from datetime import datetime

class Base_ZCP015():
    def __init__(self, export_path) -> None:
        try:
            self.path_export = export_path
            self.path_base = r'C:\Users\anderson.bones\Desktop\update-zcp015-app\files\ZCP015.xlsx'

            self.base_file_name = self.path_base.split("\\")[-1]
            self.export_file_name = self.path_export.split("\\")[-1]

            # get first day of month
            self.input_dt = datetime.today()
            self.current_date = self.input_dt.replace(day=1, hour=0, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')
        
            # get sheet names
            self.work_book_export = xl.load_workbook(self.path_export) 
            self.export_sheet=self.work_book_export.sheetnames[0]

            self.work_base = xl.load_workbook(self.path_base) 
            self.base_sheet=self.work_base.sheetnames[0]

            #read excel files
            self.df_base = pd.read_excel(self.path_base, sheet_name=self.base_sheet)
            print(f'Lendo Base: {self.base_file_name}...')

            self.df_export = pd.read_excel(self.path_export, sheet_name=self.export_sheet)
            print(f'Lendo Base: {self.export_file_name}...')

            self.dfs = []
        except Exception as e:
            print('Erro ao ler Bases!')
            

    def sort_data_pesagem(self):
        try:
            self.df_base['Hora Pesagem Inicial'] = pd.to_datetime(self.df_base['Hora Pesagem Inicial'], format='%H:%M:%S')
            self.df_base.sort_values(by=['Dt. Pesagem Inicial', 'Hora Pesagem Inicial'], inplace=True)
            print('Organizando Dt. Pesagem Inicial...')
        except Exception as e:
            print('Erro ao organizar Dt. Pesagem Inicial!')

    def remove_current_values(self):
        try:
            print('Removendo Linhas atuais...')
            self.df_base.drop(self.df_base.loc[self.df_base['Dt. Pesagem Inicial'] >= self.current_date].index, inplace=True)
            
        except Exception as e:
            print('Erro ao Remover Linhas atuais!')
        
        try:
            print('Removendo Linhas duplicadas...')
            self.new_df = self.df_base.drop_duplicates()
        except Exception as e:
            print('Erro ao remover duplicadas!')

        
    def update_data_base(self):

        try:
            print(f'Atualizando base: {self.base_file_name}')
            #self.writer = pd.ExcelWriter(self.path_base, engine='xlsxwriter') # base file
            self.dfs.append(self.df_base)
            self.dfs.append(self.df_export)
            self.df_master = pd.concat(self.dfs, axis=False)
            print("Salvando base...")
            self.df_master.to_excel('./files/result.xlsx', sheet_name=self.base_sheet, index=False, header=True)
            self.writer.close()
            print('Concluido.')
        except Exception as e:
            print(f'Erro ao autualizar base {self.base_file_name}!')
        

    def start_update(self):
        self.sort_data_pesagem()
        self.remove_current_values()
        self.update_data_base()
        
        

