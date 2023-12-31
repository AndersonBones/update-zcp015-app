import pandas as pd
from time import sleep
import os
import openpyxl as xl
from datetime import datetime
from UliPlot.XLSX import auto_adjust_xlsx_column_width
import win32com


class Base_ZCP015():
    def __init__(self, export_path) -> None:
        try:
            self.path_export = export_path
            self.rename_export_path = export_path.replace('XLSX', 'xlsx')
            os.rename(self.path_export, self.rename_export_path)
            
            self.path_base = r'F:\BIOMASSA\03. Originação\10. Controles\update-zcp015-app\ZCP015.xlsx'

            

            self.base_file_name = self.path_base.split("\\")[-1]
            print(f'Iniciando Tratamento da {self.base_file_name}')

            # get first day of month
            self.today = datetime.today()
            self.date = self.today.replace(day=1, hour=0, minute=0, second=0).strftime('%Y-%m-%d %H:%M:%S')

            self.dfs = []
        except Exception as e:
            print('Erro na inicialização!', e)
    
    def read_database(self):
        # get sheet names
        self.work_base = xl.load_workbook(self.path_base) 
        self.base_sheet=self.work_base.sheetnames[0]

        # read database
        self.df_base = pd.read_excel(self.path_base, sheet_name=self.base_sheet)
        print(f'Lendo Base: {self.base_file_name}...')
        

    
    def read_export_database(self):
        self.export_file_name = self.rename_export_path.split('/')[-1]
        # get sheet names
        self.work_book_export = xl.load_workbook(self.rename_export_path) 
        self.export_sheet=self.work_book_export.sheetnames[0]

        # read export database
        self.df_export = pd.read_excel(self.rename_export_path, sheet_name=self.export_sheet)
        print(f'Lendo Base SAP: {self.export_file_name}...')
       

    

    def sort_data_pesagem(self):
        try:
            self.df_base.sort_values(by=['Dt. Pesagem Inicial', 'Hora Pesagem Inicial'], inplace=True)
            print('Organizando Dt. Pesagem Inicial...')
        except Exception as e:
            print('Erro ao organizar Dt. Pesagem Inicial!')

    def remove_current_values(self):
        try:
            print('Removendo Linhas atuais...')
            self.df_base.drop(self.df_base.loc[self.df_base['Dt. Pesagem Inicial'] >= self.df_export.iloc[0]['Dt. Pesagem Inicial']].index, inplace=True)
            
        except Exception as e:
            print('Erro ao Remover Linhas atuais!')
        
        try:
            print('Removendo Linhas duplicadas...')
            self.new_df = self.df_base.drop_duplicates()
        except Exception as e:
            print('Erro ao remover duplicadas!')

    def auto_adjust_column(self, df):
        try:
            print('Ajustando tamanho das colunas...')
            for column in df:
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                self.writer.sheets[self.base_sheet].set_column(col_idx, col_idx, column_length+2)
        except Exception as e:
            print('Colunas ajustadas.')
        
    

    def update_data_base(self):
        try:
            print(f'Atualizando base: {self.base_file_name}')
            self.writer = pd.ExcelWriter(self.path_base, engine='xlsxwriter', date_format='d/m/yyyy') # base file
            self.dfs.append(self.df_base)
            self.dfs.append(self.df_export)
            self.df_master = pd.concat(self.dfs, axis=False)
    
            print('Base atualizada com sucesso.')
        except Exception as e:
            print(f'Erro ao autualizar base {self.base_file_name}!')
    

    def quantity_format(self):
        self.df_master['Quantidade'] = self.df_master['Quantidade'].map('{:.2f}'.format)

    def GfConvert(self):
        print('Convertendo Guia Florestal...')
        self.df_master['Guia Florestal'] = pd.to_numeric(self.df_master['Guia Florestal'], errors='coerce')
        print("Concluido.")

    def date_format(self):
        print('Ajustando formado de data...')
        self.df_master['Dt. Agendamento'] = pd.to_datetime(self.df_master['Dt. Agendamento'], format='%d %b %Y', errors='coerce').dt.date
        self.df_master['Dt. Pesagem Inicial'] = pd.to_datetime(self.df_master['Dt. Pesagem Inicial'], format='%d %b %Y', errors='coerce').dt.date
        
        self.df_master['Data Nota Fiscal'] = pd.to_datetime(self.df_master['Data Nota Fiscal'], format='%d %b %Y', errors='coerce').dt.date
        self.df_master['Data de criação'] = pd.to_datetime(self.df_master['Data de criação'], format='%d %b %Y', errors='coerce').dt.date
        
        print('Formato de data ajustado.')  

    def save_file(self):
        try:
            print("Salvando base...")
            self.df_master.to_excel(self.writer, sheet_name=self.base_sheet, index=False, header=True)
            self.auto_adjust_column(self.df_master)
            print('Salva com sucesso.')
            self.writer.close()
            
        except Exception as e:
            print(f'Erro ao salvar base {self.base_file_name}!')
    


    def start_update(self):
        self.read_database()
        self.read_export_database()
        self.sort_data_pesagem()
        self.remove_current_values()
        self.update_data_base()
        self.date_format()
        #self.quantity_format()
        self.GfConvert()
        self.save_file()

        
        

